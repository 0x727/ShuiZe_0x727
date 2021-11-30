# 调用各类插件获取子域名信息

# -*- coding:utf-8 -*-


import sys
import os
# from gevent import monkey
# monkey.patch_all()
import urllib3
import openpyxl
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from uuid import uuid4
import dns.resolver
import re
from threading import Thread
from IPy import IP
from collections import Counter
from queue import Queue
from urllib.parse import urlparse
from termcolor import cprint
from optparse import OptionParser
import os
import platform
from Plugins.saveToExcel import saveToExcel
from uuid import uuid4
import socket
import socks
import configparser
from tqdm import *
from colorama import Fore
import requests

## 调用lijiejie的子域名收集脚本
# def lijiejieSubdomain():
#     Subdomains_ips = {}             # 字典，key为子域名，value为子域名的A记录IP值
#     Subdomains_ips[domain] = []     # 主域名
#     ips_L = []
#     subdomains = ''
#     lijiejie_folder = './Plugins/infoGather/subdomain/lijiejie'
#     cprint('-' * 50 + 'Load lijiejie Subdomain ...' + '-' * 50, 'green')  # 启动lijiejie脚本
#     from Plugins.infoGather.subdomain.lijiejie.subDomainsBrute import lijiejieRun
#     # print('cd {} && python3 ./subDomainsBrute.py {}'.format(lijiejie_folder, domain))
#     # p1 = Popen('cd {} && python3 ./subDomainsBrute.py {}'.format(lijiejie_folder, domain), shell=True, stdin=PIPE, stdout=PIPE)
#     # print(p1.stdout.read().decode('gb2312'))
#     lijiejieRun(domain)
#     lijiejie_domain_file = '{}/{}.txt'.format(lijiejie_folder, domain)
#     with open(lijiejie_domain_file, 'rt') as f:
#         for each_line in f.readlines():
#             each_line_split = each_line.split('\t')
#             subdomain = each_line_split[0].strip()                  # 子域名
#             ips = each_line_split[1].strip('\n')                    # 子域名的dns解析A记录IP
#             # print(subdomain, ips)
#             for ip in ips.split(','):
#                 ips_L.append(ip.strip())
#             # print(subdomain, ips_L)
#             Subdomains_ips[subdomain] = ips_L
#             ips_L = []
#
#     os.remove(lijiejie_domain_file)                         # 删除临时文件
#
#     lijiejie_tmp = lijiejie_folder + '/tmp'                 # 删除tmp目录
#     if os.path.isdir(lijiejie_tmp):
#         shutil.rmtree(lijiejie_tmp, True)
#
#     return Subdomains_ips

# 进度条


# 判断是否是IP
def isIP(str):
    p = re.compile('^((25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(25[0-5]|2[0-4]\d|[01]?\d\d?)$')
    if p.match(str):
        return True
    else:
        return False

# 获取github敏感信息
def get_GitSensitiveInfo(github_txt, raw_url_emails):
    cf = configparser.ConfigParser()
    cf.read("./iniFile/config.ini")
    secs = cf.sections()
    github_keywords = eval(cf.get('github keywords', 'github_keywords'))

    line_urls = {}
    gitSensitiveInfo = []

    with open(github_txt, 'rt') as f:
        content = f.readlines()
        for line, each in enumerate(content):
            if '[------------------] ' in each:
                line_urls[str(line + 1)] = each.split('[------------------] ')[1]

    lines = list(line_urls.keys())

    # print(line_urls)
    def get_githubAddr(line):
        for i, num in enumerate(lines):
            # print(line)
            if i < len(lines) - 1:
                # print(line, int(num), int(lines[i + 1]))
                if int(num) <= line <= int(lines[i + 1]):
                    return int(num)
            elif line > int(lines[-1]):
                return int(lines[-1])

    for keyword in github_keywords:
        for line, each in enumerate(content):
            if line < len(content) - 1:
                if keyword in each:
                    # print(line)
                    githubAddr = get_githubAddr(line)
                    # print(githubAddr)
                    if githubAddr:
                        raw_url = content[int(githubAddr) - 1].replace('[------------------]', '').strip()
                        try:
                            emails = str(raw_url_emails[raw_url])
                            print('github address: [line:{}] {}'.format(githubAddr, raw_url))
                            print('[emails] : {}'.format(emails))
                            print('[{}] [line:{}] {}'.format(keyword, line, content[line - 1].strip()))
                            print('[{}] [line:{}] {}'.format(keyword, line + 1, content[line].strip()))
                            print('[{}] [line:{}] {}'.format(keyword, line + 2, content[line + 1].strip()))
                            gitSensitiveInfo.append(['gitAddress', githubAddr, raw_url, emails])
                            gitSensitiveInfo.append([keyword, line, content[line - 1].strip(), emails])
                            gitSensitiveInfo.append([keyword, line + 1, content[line].strip(), emails])
                            gitSensitiveInfo.append([keyword, line + 2, content[line + 1].strip(), emails])
                            gitSensitiveInfo.append(['-' * 50, '-' * 50, '-' * 50, '-' * 50])
                        except Exception as e:
                            pass

    return gitSensitiveInfo

# 打印脚本跑出了几个新的子域名，并返回最新最全的子域名列表  传递两个列表，old是前面收集好的子域名，new是刚跑完的脚本收集的子域名，进行比较.
def printGetNewSubdomains(old_subdomains, new_subdomains):
    if len(old_subdomains) > 0:
        newSubdomains = list(set(new_subdomains) - set(old_subdomains))
        print('[new :{}] {}'.format(len(newSubdomains), newSubdomains))
    return list(set(new_subdomains + old_subdomains))


# subdomains3脚本调用
def subdomains3():
    cprint('-' * 50 + 'Load subdomains3 ...' + '-' * 50, 'green')
    from Plugins.infoGather.subdomain.subdomain3.brutedns import run_subdomains
    Subdomains_ips = run_subdomains(domain)
    return Subdomains_ips


# dns域传送漏洞
def dnsZoneTransfer():
    pass


# 从fofa收集代理
def getSocksProxy():
    cprint('-' * 50 + 'Load getSocksProxy ...' + '-' * 50, 'green')
    from Plugins.infoGather.SocksProxy.getSocksProxy import run_getSocksProxy
    socksProxysDict = run_getSocksProxy()

    # 保存到excel
    # socksProxysSheet = saveToExcel(excelSavePath, excel, '代理')
    # socksProxysSheet.saveSocksProxys(socksProxysDict)

# 备案反查顶级域名
def beian2NewDomain():
    cprint('-' * 50 + 'Load beian2NewDomain ...' + '-' * 50, 'green')
    from Plugins.infoGather.subdomain.beian2NewDomain.beian2domain import run_beian2domain
    beianNewDomains, companyName = run_beian2domain(domain)

    for _ in beianNewDomains:
        newDomains.append(_[1])

    # 保存到excel
    beianNewDomainsSheet = saveToExcel(excelSavePath, excel, '备案反查顶级域名')
    beianNewDomainsSheet.saveBeianNewDomains(beianNewDomains)
    return companyName


# 从爱企查获取目标相关信息
def Aiqicha(companyName):
    cprint('-' * 50 + 'Load Aiqicha ...' + '-' * 50, 'green')
    if not companyName:
        return

    cprint("查询【{}】公司架构".format(companyName), 'red')

    from Plugins.infoGather.subdomain.Aiqicha.Aiqicha import run_aiqicha
    selfIcpinfo_infos, invest_infos, holds_infos, branch_infos = run_aiqicha(companyName)

    # 保存到excel
    aiqichaSheet = saveToExcel(excelSavePath, excel, '爱企查')
    aiqichaSheet.saveAiqicha(selfIcpinfo_infos, invest_infos, holds_infos, branch_infos)



# 判断是否是泛解析
def checkPanAnalysis():
    cprint('-' * 50 + 'check Pan-Analysis ...' + '-' * 50, 'green')
    panDomain = 'sadfsadnxzjlkcxjvlkasdfasdf.{}'.format(domain)
    try:
        dns_A_ips = [j for i in dns.resolver.query(panDomain, 'A').response.answer for j in i.items]
        print(dns_A_ips)
        cprint('[泛解析] {} -> {}'.format(panDomain, dns_A_ips), 'red')
        return True
    except Exception as e:
        cprint('[不是泛解析] :{}'.format(e.args), 'red')
        return False



# 调用kSubdomain脚本
def callKsubdomain():
    cprint('-' * 50 + 'Load ksubdomain ...' + '-' * 50, 'green')
    from Plugins.infoGather.subdomain.ksubdomain.ksubdomain import run_ksubdomain
    ksubdomains = run_ksubdomain(domain)
    return ksubdomains




# theHarvest脚本调用
def theHarvester():
    cprint('-' * 50 + 'Load theHarvest ...' + '-' * 50, 'green')
    from Plugins.infoGather.subdomain.theHarvester.theHarvester import run_theHarvester
    theHarvesterIp, emails, hosts = run_theHarvester(domain)
    print(hosts)
    theHarvesterSubdomains = []
    subdomain = None
    for host in list(set(hosts)):
        if '/' not in host and ' ' not in host:
            domain_ip = host.strip().split(':')
            if len(domain_ip) == 2:
                subdomain, ip = [domain_ip[0]], domain_ip[1]
            elif len(domain_ip) == 1:
                subdomain, ip = domain_ip, None
        if subdomain:
            theHarvesterSubdomains.extend(subdomain)

    # 测试

    # 检测邮箱的真实性
    cprint('-' * 50 + 'Load verifyEmails ...' + '-' * 50, 'green')
    from Plugins.infoGather.subdomain.verifyEmails.VerifyEmails import run_verifyEmails
    aliveEmails = run_verifyEmails(emails)

    # 保存到excel
    theHarvesterIpSheet = saveToExcel(excelSavePath, excel, 'theHarvester—IP')
    theHarvesterIpSheet.saveTheHarvesterIp(theHarvesterIp)

    emailsSheet = saveToExcel(excelSavePath, excel, '邮箱')
    emailsSheet.saveEmails(emails, aliveEmails)
    return list(set(theHarvesterSubdomains))


# 调用virustotal|ce.baidu.com|www.threatcrowd.org|url.fht.im|的子域名收集脚本
def othersApiSubdomain():
    cprint('-' * 50 + 'Load VirusTotal threatcrowd url.fht.im ...' + '-' * 50, 'green')
    from Plugins.infoGather.subdomain.othersApiSubdomains.othersApiSubdomains import othersApiRun
    othersApiTotalSubdomains = othersApiRun(domain)          # 列表，存放子域名
    return othersApiTotalSubdomains

# 调用github api的子域名收集脚本
def githubApiSubdomain():
    cprint('-' * 50 + 'Load Github Api Subdomain ...' + '-' * 50, 'green')
    from Plugins.infoGather.subdomain.githubSubdomains.githubSubdomains import githubApiRun
    githubApiSubdomains, raw_url_emails = githubApiRun(domain, save_fold_path)          # 列表，存放子域名

    # 保存到excel
    githubSheet = saveToExcel(excelSavePath, excel, 'Github敏感信息')
    github_txt = r'{}/{}_github.txt'.format(save_fold_path, domain)
    if os.path.exists(github_txt):
        gitSensitiveInfo = get_GitSensitiveInfo(github_txt, raw_url_emails)
        githubSheet.saveGithub(gitSensitiveInfo)

    return githubApiSubdomains


# 调用Sublist3r子域名收集脚本
def Sublist3r():
    print('[+] Load Sublist3r Subdomain ...')
    from Plugins.infoGather.subdomain.Sublist3r.sublist3r import sublist3rRun
    sublist3rSubdomains = sublist3rRun(domain)
    return sublist3rSubdomains

# 调用爬虫
def SpiderSubdomain():
    cprint('-' * 50 + 'Load Spider ...' + '-' * 50, 'green')  # 启动百度爬虫
    spiderSheet = saveToExcel(excelSavePath, excel, '爬虫')

    # 百度爬虫
    def BaiduSubdomain():
        cprint('Load BaiduSpider ...', 'green')  # 启动百度爬虫
        from Plugins.infoGather.subdomain.Spider.Baidu.baidu import BaiduSpider
        bdSubdomains, links = BaiduSpider().run_subdomain(domain)
        # 保存到excel
        spiderSheet.saveSpider('百度', links)
        return bdSubdomains

    # 必应爬虫
    def BingSubdomain():
        cprint('Load BingSpider ...', 'green')  # 启动必应爬虫
        from Plugins.infoGather.subdomain.Spider.Bing.bing import BingSpider
        bingSubdomains, links = BingSpider().run_subdomain(domain)
        # 保存到excel
        spiderSheet.saveSpider('必应', links)
        return bingSubdomains

    bdSubdomains = BaiduSubdomain()
    bingSubdomains = BingSubdomain()
    spiderSubdomains = list(set(bdSubdomains + bingSubdomains))

    return spiderSubdomains

# 抓取https域名的证书dns信息
def crawlCerts(subdomains):
    cprint('-' * 50 + 'Load crawlCerts ...' + '-' * 50, 'green')  # 启动证书爬虫
    from Plugins.infoGather.subdomain.Certs.crawlCerts import crawlCerts
    certsSubdomains, trustedDomainDict, _newDomains = crawlCerts(domain, subdomains).run()

    newDomains.extend(_newDomains)
    # 保存到excel
    certSheet = saveToExcel(excelSavePath, excel, '证书')
    certSheet.saveCert(trustedDomainDict)

    return certsSubdomains

# 调用友链爬虫
def FriendChinsSubdomain(temp_subdomains):
    cprint('-' * 50 + 'Load FriendChins ...' + '-' * 50, 'green')  # 启动友链爬虫
    from Plugins.infoGather.subdomain.FriendChins.crawlFriendChins import FriendChins
    fcSubdomains = FriendChins(domain, temp_subdomains).run()
    return fcSubdomains


# 整合子域名，对所有子域名解析A记录
def checkCDN_queryA_subdomains(Subdomains_ips, subdomains):
    cprint('-' * 50 + 'check subdomains CDN and query ip ...' + '-' * 50, 'green')  # 整合所有子域名
    tmp_subdomains = []
    for subdomain in subdomains:
        if '.{}'.format(domain) in subdomain:
            tmp_subdomains.append(subdomain)
    subdomains = list(set(tmp_subdomains))


    print('Check CDN [{}] subdomains'.format(len(subdomains)))
    from Plugins.infoGather.subdomain.CDN import checkCDN
    notCDNSubdomains, CDNSubdomainsDict = checkCDN.run_checkCDN(subdomains)

    print('Query the A record of [{}] subdomains'.format(len(subdomains)))
    from Plugins.infoGather.subdomain.queryA import queryA
    Subdomains_ips = queryA.run_queryA(Subdomains_ips, subdomains)

    # 保存到excel
    queryASheet = saveToExcel(excelSavePath, excel, '子域名A记录')
    queryASheet.saveQueryA(Subdomains_ips, CDNSubdomainsDict)

    return Subdomains_ips, CDNSubdomainsDict, notCDNSubdomains

# host碰撞
def hostCollide(Subdomains_ips):
    cprint('-' * 50 + 'run_hostCollide ...' + '-' * 50, 'green')  # 启动网络空间引擎
    from Plugins.infoGather.subdomain.hostCollide import hostCollide
    hostCollideResult, censysIPS = hostCollide.run_hostCollide(domain, Subdomains_ips)

    # 保存到excel
    queryASheet = saveToExcel(excelSavePath, excel, 'HOST碰撞')
    queryASheet.saveHostCollide(hostCollideResult)

    return censysIPS

# 获取所有子域名的参数链接和后台链接（存活）
def run_ParamLinks():
    cprint('-' * 50 + 'run_ParamLinks ...' + '-' * 50, 'green')  # 启动网络空间引擎
    from Plugins.infoGather.ParamSpider.paramSpider import getParamLinks
    paramLinks, htLinks = getParamLinks(domain)

    # 保存到excel
    paramHtLinksSheet = saveToExcel(excelSavePath, excel, '动态链接和后台地址')
    paramHtLinksSheet.saveparamHtLinks(paramLinks, htLinks)

    # 如果动态链接的个数大于1000，
    if len(paramLinks) > 1000:
        paramLinks = []

    return paramLinks

# 整理IP，获取C段IP
def get_CIP(Subdomains_ips, CDNSubdomainsDict, censysIPS):
    cprint('-' * 50 + 'get_CIP ...' + '-' * 50, 'green')  # 整理C段IP
    # 过滤内网IP
    def is_internal_ip(ip_subnet):
        ip_subnet_list = ip_subnet.split('.')
        if ip_subnet_list[0] == '10' or ip_subnet_list[0] == '127':
            return True
        elif ip_subnet_list[0] == '172' and 15 < int(ip_subnet_list[1]) < 32:
            return True
        elif ip_subnet_list[0] == '192' and ip_subnet_list[1] == '168':
            return True
        else:
            return False

    ips = []
    CIP_List = []
    CIP_List_all = []

    for subdomain in Subdomains_ips:
        if CDNSubdomainsDict[subdomain] == 'NOT':           # 如果该子域名没有CDN，则开始统计解析出来的IP
            ip_List = Subdomains_ips[subdomain]
            for ip in ip_List:
                if not is_internal_ip(ip):
                    ips.append(ip)

    ips.extend(censysIPS)

    for ip in list(set(ips)):
        c_subnet = str(IP(ip).make_net('255.255.255.0')).rsplit('.', 1)[0] + '.0'
        CIP_List_all.append(c_subnet)

    global ip_count
    ip_count = Counter(CIP_List_all)
    cprint(ip_count, 'red')
    import configparser
    cf = configparser.ConfigParser()
    cf.read("./iniFile/config.ini")
    c_nums = cf.get('C nums', 'c_nums')

    for ip in ip_count:
        if ip_count[ip] > int(c_nums):
            CIP_List.append(ip)

    return CIP_List
    # return list(set(CIP_List))


# 调用网络空间引擎,查询根域名和C段IP的资产
def run_webSpace(domain, SubdomainAndNotCDNIPs, CIP_List, fofaTitle):
    cprint('-' * 50 + 'run_webSpace ...' + '-' * 50, 'green')  # 启动网络空间引擎
    from Plugins.infoGather.WebspaceSearchEngine import fofaApi, shodanApi, quakeApi, qianxinApi
    webSpaceSheet = saveToExcel(excelSavePath, excel, '网络空间搜索引擎')
    serviceSheet = saveToExcel(excelSavePath, excel, '服务')

    webSpace_web_host_port = []         # 存放开放web服务
    webSpace_service_host_port = []     # 存放除了Web的其他服务

    # fofa搜索引擎信息收集
    def run_fofa():
        # 查询域名
        if domain:
            query_str = 'domain="{}"'.format(domain)
            fofa_Results, fofa_web_host_port, fofa_service_host_port = fofaApi.query_domain(query_str)
            if fofa_Results:
                webSpaceSheet.saveWebSpace('fofa', fofa_Results, query_str) # 将网络空间搜索引擎的结果保存到webSpace项里
                # save_webSpace(fofa_Results, 'fofa', query_str)
                webSpace_web_host_port.extend(fofa_web_host_port)
                webSpace_service_host_port.extend(fofa_service_host_port)

        # 查询C段IP
        if CIP_List:
            for c_subnet in CIP_List:
                query_str = 'ip="{}/24"'.format(c_subnet)
                fofa_Results, fofa_web_host_port, fofa_service_host_port = fofaApi.query_ip(query_str)
                if fofa_Results:
                    webSpaceSheet.saveWebSpace('fofa', fofa_Results, query_str)
                    webSpace_web_host_port.extend(fofa_web_host_port)
                    webSpace_service_host_port.extend(fofa_service_host_port)

        if fofaTitle:
            query_str = 'title="{}" && country="CN"'.format(fofaTitle)
            fofa_Results, fofa_web_host_port, fofa_service_host_port = fofaApi.query_domain(query_str)
            if fofa_Results:
                webSpaceSheet.saveWebSpace('fofa', fofa_Results, query_str) # 将网络空间搜索引擎的结果保存到webSpace项里
                # save_webSpace(fofa_Results, 'fofa', query_str)
                webSpace_web_host_port.extend(fofa_web_host_port)
                webSpace_service_host_port.extend(fofa_service_host_port)



    # shodan搜索引擎信息收集
    def run_shodan():
        # 查询域名
        if domain:
            query_str = 'hostname:"{}"'.format(domain)
            shodan_Results, shodan_web_host_port, shodan_service_host_port = shodanApi.query_domain(query_str)
            if shodan_Results:
                webSpaceSheet.saveWebSpace('shodan', shodan_Results, query_str)
                webSpace_web_host_port.extend(shodan_web_host_port)
                webSpace_service_host_port.extend(shodan_service_host_port)

        # 查询C段IP
        if CIP_List:
            for c_subnet in CIP_List:
                query_str = 'net:"{}/24"'.format(c_subnet)
                shodan_Results, shodan_web_host_port, shodan_service_host_port = shodanApi.query_ip(query_str)
                if shodan_Results:
                    webSpaceSheet.saveWebSpace('shodan', shodan_Results, query_str)
                    webSpace_web_host_port.extend(shodan_web_host_port)
                    webSpace_service_host_port.extend(shodan_service_host_port)

    # quake搜索引擎信息收集
    def run_quake():
        # 查询域名
        if domain:
            query_str = 'domain:"{}" AND country:"China"'.format(domain)
            quake_Results, quake_web_host_port, quake_service_host_port = quakeApi.query_domain(query_str)
            if quake_Results:
                webSpaceSheet.saveWebSpace('quake', quake_Results, query_str)
                webSpace_web_host_port.extend(quake_web_host_port)
                webSpace_service_host_port.extend(quake_service_host_port)

        # 查询C段IP
        if CIP_List:
            for c_subnet in CIP_List:
                query_str = 'ip:"{}/24"'.format(c_subnet)
                quake_Results, quake_web_host_port, quake_service_host_port = quakeApi.query_ip(query_str)
                if quake_Results:
                    webSpaceSheet.saveWebSpace('quake', quake_Results, query_str)
                    webSpace_web_host_port.extend(quake_web_host_port)
                    webSpace_service_host_port.extend(quake_service_host_port)

        if fofaTitle:
            query_str = 'title:"{}" AND country:"China"'.format(fofaTitle)
            quake_Results, quake_web_host_port, quake_service_host_port = quakeApi.query_ip(query_str)
            if quake_Results:
                webSpaceSheet.saveWebSpace('quake', quake_Results, query_str)
                webSpace_web_host_port.extend(quake_web_host_port)
                webSpace_service_host_port.extend(quake_service_host_port)

    # qianxin搜索引擎信息收集
    def run_qianxin():
        # 查询域名
        if domain:
            query_str = '(domain="{}")&&(country=="中国")'.format(domain)
            qianxin_Results, qianxin_web_host_port, qianxin_service_host_port = qianxinApi.query_domain(query_str)
            if qianxin_Results:
                webSpaceSheet.saveWebSpace('qianxin', qianxin_Results, query_str)
                webSpace_web_host_port.extend(qianxin_web_host_port)
                webSpace_service_host_port.extend(qianxin_service_host_port)

        # 查询C段IP
        if CIP_List:
            for c_subnet in CIP_List:
                query_str = 'ip="{}/24"'.format(c_subnet)
                qianxin_Results, qianxin_web_host_port, qianxin_service_host_port = qianxinApi.query_ip(query_str)
                if qianxin_Results:
                    webSpaceSheet.saveWebSpace('qianxin', qianxin_Results, query_str)
                    webSpace_web_host_port.extend(qianxin_web_host_port)
                    webSpace_service_host_port.extend(qianxin_service_host_port)

        if fofaTitle:
            query_str = '(title="{}")&&(country=="中国")'.format(fofaTitle)
            qianxin_Results, qianxin_web_host_port, qianxin_service_host_port = qianxinApi.query_ip(query_str)
            if qianxin_Results:
                webSpaceSheet.saveWebSpace('qianxin', qianxin_Results, query_str)
                webSpace_web_host_port.extend(qianxin_web_host_port)
                webSpace_service_host_port.extend(qianxin_service_host_port)


    # 对子域名和非CDN的IP进行fofa查询
    def run_fofaOne(subdomainAndIP_Q):
        while not subdomainAndIP_Q.empty():
            subdomainOrIp = subdomainAndIP_Q.get()
            if isIP(subdomainOrIp):
                query_str = 'ip="{}"'.format(subdomainOrIp)
            else:
                query_str = 'domain="{}"'.format(subdomainOrIp)
            fofa_Results, fofa_web_host_port, fofa_service_host_port = fofaApi.query_ip(query_str)
            if fofa_Results:
                webSpaceSheet.saveWebSpace('fofa', fofa_Results, query_str)  # 将网络空间搜索引擎的结果保存到webSpace项里
                # save_webSpace(fofa_Results, 'fofa', query_str)
                webSpace_web_host_port.extend(fofa_web_host_port)
                webSpace_service_host_port.extend(fofa_service_host_port)

    run_fofa()
    run_shodan()
    run_quake()
    run_qianxin()


    # fofa跑所有子域名解析出来的IP
    if SubdomainAndNotCDNIPs:
        subdomainAndIP_Q = Queue(-1)
        for subdomainAndIP in SubdomainAndNotCDNIPs:
            subdomainAndIP_Q.put(subdomainAndIP)
        threads = []
        for t_id in range(5):
            t = Thread(target=run_fofaOne, args=(subdomainAndIP_Q, ))
            threads.append(t)
            t.start()
        for t in threads:
            t.join()

    serviceResult = []
    for _ in webSpace_service_host_port:            # 去重
        if _ not in serviceResult:
            serviceResult.append(_)

    webSpace_service_host_port = serviceResult
    # 将非Web服务的结果保存到service项里
    serviceSheet.saveService(webSpace_service_host_port)


    return webSpace_web_host_port, webSpace_service_host_port


# 整理fofaTitle结果的域名和IP
def collation_fofaDomainIP(webSpace_web_host_port):
    ips = []
    fofaTitle_IPs = []
    fofaTitle_newDomains = []

    for _ in webSpace_web_host_port:
        a = urlparse(_)
        if a.scheme:
            newdomain_ip = a.netloc.split(':')[0]
        else:
            newdomain_ip = a.path.split(':')[0]
        if isIP(newdomain_ip):
            ips.append(newdomain_ip)
        else:
            fofaTitle_newDomains.append(newdomain_ip)

    for ip in list(set(ips)):
        ip_C = str(IP(ip).make_net('255.255.255.0')).rsplit('.', 1)[0] + '.0'
        fofaTitle_IPs.append(ip_C)

    global ip_count
    ip_count = Counter(fofaTitle_IPs)
    newDomains.extend(fofaTitle_newDomains)

# ip反查域名，并将域名结果保存到Subdomains_ips列表里，并且存放到ip2domain_dict里
def get_ip2domain():
    cprint('-' * 50 + 'ip to domain ...' + '-' * 50, 'green')  # 对IP进行域名反查
    from Plugins.infoGather.subdomain.ip2domain import getIp2Domain

    ip2domain_dict, _newDomains = getIp2Domain.run_ip2domain(domain, allTargets_Queue)  # ip2domain_dict字典，key为IP，value为反查的域名

    # 和目标关联的相关域名
    newDomains.extend(_newDomains)

    # 去重
    ip2domainSubdomains = []                        # 反查出来的子域名列表    ['ca.hbu.edu.cn', 'jwjcc.bdu.edu.cn', 'yzuuc.hbu.cn']
    for subdomains in ip2domain_dict.values():
        for subdomain in subdomains:
            if domain:
                if domain in subdomain:
                    ip2domainSubdomains.append(subdomain)
            else:
                ip2domainSubdomains.append(subdomain)
    ip2domainSubdomains = list(set(ip2domainSubdomains))

    ip2domainSheet = saveToExcel(excelSavePath, excel, 'ip反查域名')    # 创建一个ip反查域名页
    ip2domainSheet.saveIp2Domain(ip2domain_dict)

    return ip2domainSubdomains      # 返回ip反查得到的域名列表


# 对IP进行归属地查询
def get_ip_address(web_ip_list):
    cprint('-' * 50 + 'get ip address ...' + '-' * 50, 'green')  # 对IP进行归属地查询
    from Plugins.infoGather.subdomain.ipAddress import getIpAddress
    ip_address_dict = getIpAddress.run_getIpAddress(web_ip_list)            # 字典，key为IP，value为归属地
    return ip_address_dict


# 整理开放web服务的host
def collation_web_host(Subdomains_ips, webSpace_web_host_port, ip2domainSubdomains):
    cprint('-' * 50 + 'collation_web_host ...' + '-' * 50, 'green')  # 启动web服务收集
    web_host_port = []                      # 存放最终的开放web服务的host
    web_host_port_temp = []                 # web_host_port临时列表
    web_ip_list = []                        # 存放开放web服务的IP


    for subdomain in list(set(list(Subdomains_ips.keys()) + ip2domainSubdomains)):
        if ':' in subdomain:                        # ip2domainSubdomains的结果里有些类似于221.192.236.146:999这种结果，所以不加80端口
            web_host_port_temp.append(subdomain)
        else:
            web_host_port_temp.append('{}:80'.format(subdomain))

    web_host_port_temp.extend(webSpace_web_host_port)
    # print('[{}] {}'.format(len(web_host_port), web_host_port))
    web_host_port_temp = list(set(web_host_port_temp))
    # print('[{}] {}'.format(len(web_host_port), web_host_port))

    # 整合url，格式规范。全部是http(s)://www.domain.com:xxxx
    for host_port in web_host_port_temp:
        host_port_urlparse = urlparse(host_port)
        if not host_port_urlparse.scheme:           # 如果没有http（https）， 则加上。如果是443端口，则加https，其他端口加http
            if ':' in host_port:
                try:
                    host, port = host_port.split(':')
                    if isIP(host):
                        web_ip_list.append(host)
                    if port == '443':
                        host_port = 'https://{}'.format(host)
                    elif port == '80':
                        host_port = 'http://{}'.format(host)
                    else:
                        host_port = 'http://{}'.format(host_port)
                except Exception as e:
                    pass
            else:
                host_port = 'http://{}'.format(host_port)
        else:   # 如果有https或者http，则不加
            host_port = host_port
        web_host_port.append(host_port)

    web_host_port = list(set(web_host_port))    # 去重
    web_ip_list = list(set(web_ip_list))

    return web_host_port, web_ip_list


# 内网端口扫描
def scan_IntranetPorts():


    from Plugins.infoGather.Intranet.scanPort import scanPort
    from Plugins.infoGather.Intranet import getMoreIp

    tqdm.write(Fore.BLACK + '-' * 50 + 'scan_IntranetPorts ...' + '-' * 50)
    web_host_port, service_host_port, alive_host_List = scanPort.run_ScanPort(allTargets_Queue, proxy)

    tqdm.write(Fore.BLACK + '-' * 50 + 'get_IntranetHostName and IP ...' + '-' * 50)
    alive_hostname_ips = getMoreIp.run_getMoreIp(alive_host_List)        # 通过oxid获取主机名和更多的内网IP

    # 写入表格里
    intranetServiceSheet = saveToExcel(excelSavePath, excel, '内网服务')
    intranetServiceSheet.saveService(service_host_port)

    intranetHostNameIpsSheet = saveToExcel(excelSavePath, excel, '内网主机名和IP')
    intranetHostNameIpsSheet.saveHostNameAndIps(alive_hostname_ips)

    return web_host_port, alive_host_List

# 筛选存活并获取标题
def run_getWebTitle(web_host_port, ip_address_dict):
    tqdm.write(Fore.BLACK + '-' * 50 + 'run_getWebTitle ...' + '-' * 50)  # 筛选存活并获取标题
    from Plugins.infoGather.webInfo import getWebTitle
    if isIntranet == 1:
        threadNum = 10        # 如果是扫内网，则线程为5
    else:
        threadNum = 300      # 扫外网则线程为300
    web_Titles = getWebTitle.run_getWebTitle(web_host_port, ip_address_dict, requests_proxies, threadNum)
    # print(web_Titles)
    alive_Web = []  # 存活的web服务
    for each in web_Titles:
        if each[1] != 65535:
            alive_Web.append(each[0])

    # 写入表格里
    webTitileSheet = saveToExcel(excelSavePath, excel, '存活网站标题')  # 创建一个ip反查域名页
    webTitileSheet.saveWebTitle(web_Titles)

    return web_Titles, alive_Web


# Web漏洞检测
def detect_webVul(alive_Web):

    tqdm.write(Fore.BLACK + '-' * 50 + 'detect Web vul' + '-' * 50)  # 探测各种漏洞
    webVul_list = []  # 存储Web漏洞，每个元素都是一个列表。[['shiro', 'http://127.0.0.1'], ['weblogic', 'http://127.0.0.1'], ['phpstudy', 'http://127.0.0.1']]

    vul_path = os.getcwd() + '/Plugins/Vul/Web/'
    sys.path.append(vul_path)  # 添加环境变量
    vulList = filter(lambda x: (True, False)[x[-3:] == 'pyc' or x[-5:] == '__.py' or x[:2] == '__'],
                     os.listdir(vul_path))  # 获取漏洞脚本

    # 内网跑的漏洞
    intPassVul = ['Jboss.py', 'phpstudy.py', 'weblogic.py', 'cms.py', 'yongyou.py', 'easyConnect.py', 'shiro.py']

    for vulName in vulList:
        tqdm.write(Fore.BLACK + '-' * 50 + 'detect ' + vulName[:-3] + '-' * 50)  # 探测各种漏洞
        md = __import__(vulName[:-3])  # 导入类
        try:
            if hasattr(md, 'Detect'):
                detect = getattr(md, 'Detect')  # 获取类

                alive_Web_queue = Queue(-1)  # 将存活的web存入队列里
                for _ in alive_Web:
                    alive_Web_queue.put(_)

                threads = []
                if isIntranet == 1:
                    threadNum = 30  # 如果是扫内网，则线程为5
                    if vulName in intPassVul:
                        pass
                    else:
                        tqdm.write(Fore.BLACK + '内网不跑{}漏洞'.format(vulName))
                        continue
                else:
                    threadNum = 100  # 扫外网则线程为300

                pbar = tqdm(total=alive_Web_queue.qsize(), desc="检测Web漏洞", ncols=150)  # total是总数

                for num in range(1, threadNum + 1):
                    t = detect(alive_Web_queue, pbar, webVul_list, requests_proxies)  # 实例化漏洞类，传递参数：存活web的队列，  存储漏洞的列表
                    threads.append(t)
                    t.start()
                for t in threads:
                    t.join()

                pbar.close()
        except Exception as e:
            tqdm.write(Fore.BLACK + r'[-] Load Vul [{}] Error: {}'.format(vulName, e.args))
            continue

    return webVul_list

# 参数漏洞检测
def detect_paramVul(param_Links):
    tqdm.write(Fore.BLACK + '-' * 50 + 'detect param vul' + '-' * 50)  # 探测各种参数漏洞-注入
    paramVul_list = []  # 存储参数漏洞，每个元素都是一个列表。[['SQL', 'http://127.0.0.1/a.php?id=1'], ['SQL', 'http://127.0.0.1/a.php?id=2']]

    vul_path = os.getcwd() + '/Plugins/Vul/Param/'
    sys.path.append(vul_path)  # 添加环境变量
    vulList = filter(lambda x: (True, False)[x[-3:] == 'pyc' or x[-5:] == '__.py' or x[:2] == '__'],
                     os.listdir(vul_path))  # 获取漏洞脚本


    for vulName in vulList:
        tqdm.write(Fore.BLACK + '-' * 50 + 'detect ' + vulName[:-3] + '-' * 50)  # 探测各种漏洞
        md = __import__(vulName[:-3])  # 导入类
        try:
            paramVul_list = md.detect(param_Links)
        except Exception as e:
            tqdm.write(Fore.BLACK + r'[-] Load Vul [{}] Error: {}'.format(vulName, e.args))
            continue

    return paramVul_list



# 未授权漏洞检测
def detect_unauthWeakVul(service_host_port):
    tqdm.write(Fore.BLACK + '-' * 50 + 'detect unauth vul' + '-' * 50)  # 探测各种漏洞
    tqdm.write(Fore.BLACK + 'service_host_port : {}'.format(service_host_port))

    service_host_port_queue = Queue(-1)            # 队列
    for _ in service_host_port:
        service_host_port_queue.put((_))

    # 键值队，key为漏洞的名字，value为漏洞插件名
    serviceVulName = {'redis': 'unAuthRedis', 'elastic': 'unAuthElastic', 'mongodb': 'unAuthMongodb', 'ldaps': 'unAuthLdaps',
                      'zookeeper': 'unAuthZookeeper', 'ftp': 'weakFTP', 'ssh': 'weakSSH', 'mssql': 'weakMSSQL',
                      'mysql': 'weakMYSQL', 'rdp': 'weakRDP'}

    # 弱口令漏洞-密码字典文件地址
    weakTxtDict = {'ftp': 'dic_password_ftp.txt', 'ssh': 'dic_password_ssh.txt', 'mssql': 'dic_password_sqlserver.txt',
                      'mysql': 'dic_password_mysql.txt', 'rdp': 'dic_password_rdp.txt'}

    # 存放弱口令密码
    serviceWeakPwds = {}
    # 读取密码字典
    for protocol in weakTxtDict.keys():
        weakPwdTxt = './iniFile/PwdTxt/{}'.format(weakTxtDict[protocol])
        with open(weakPwdTxt, 'rt') as f:
            serviceWeakPwds[protocol] = f.readlines()

    unauthWeakVul_list = []  # 存储未授权漏洞，每个元素都是一个列表。[['redis', 'http://127.0.0.1'], ['elastic', 'http://127.0.0.1']]

    unauthVul_path = os.getcwd() + '/Plugins/Vul/Service/'
    sys.path.append(unauthVul_path)  # 添加环境变量

    # 多线程探测未授权-弱口令漏洞
    def detect_unauthWeak(protocol, ip, port):
        if protocol in serviceVulName.keys():
            vulName = serviceVulName[protocol]
            # 跑域名和C段的时候默认要跑弱口令
            # if not domain and not cSubnet:
            if not weak and protocol in weakTxtDict.keys():
                return
            if protocol in serviceWeakPwds.keys():
                weakPwdsList = serviceWeakPwds[protocol]        # 弱口令密码列表
            else:
                weakPwdsList = []

            tqdm.write(Fore.BLACK + 'test [{}] : {} {}'.format(vulName, ip, port))

            try:
                md = __import__(vulName)  # 导入类
                if hasattr(md, 'Detect'):
                    detect = getattr(md, 'Detect')  # 获取类
                    detect(ip, port, unauthWeakVul_list).run_detect(weakPwdsList)
            except Exception as e:
                tqdm.write(Fore.BLACK + r'[-] Load Vul [{}] Error: {}'.format(vulName, e.args))

    pbar = tqdm(total=len(service_host_port), desc="检测未授权漏洞", ncols=150)  # total是总数
    for _ in service_host_port:
        protocol, ip, port = _
        detect_unauthWeak(protocol, ip, port)
        pbar.update(1)
    pbar.close()  # 关闭进度条



    return unauthWeakVul_list



# Windows漏洞检测
def detect_winVul(alive_host_List):
    cprint('-' * 50 + 'detect Windows vul' + '-' * 50, 'green')  # 探测Windows漏洞
    winVul_list = []  # 存储Windows漏洞，每个元素都是一个列表。[['CVE-2020-0796', '127.0.0.1'], ['MS17010', '127.0.0.1']]

    vul_path = os.getcwd() + '/Plugins/Vul/Win/'
    sys.path.append(vul_path)  # 添加环境变量
    vulList = filter(lambda x: (True, False)[x[-3:] == 'pyc' or x[-5:] == '__.py' or x[:2] == '__'],
                     os.listdir(vul_path))  # 获取漏洞脚本

    for vulName in vulList:
        cprint('-' * 50 + 'detect ' + vulName[:-3] + '-' * 50, 'green')  # 探测各种漏洞
        md = __import__(vulName[:-3])  # 导入类
        try:
            if hasattr(md, 'Detect'):
                detect = getattr(md, 'Detect')  # 获取类

                alive_host_queue = Queue(-1)  # 将存活的主机存入队列里
                for _ in alive_host_List:
                    alive_host_queue.put(_)

                threads = []
                if isIntranet == 1:
                    threadNum = 5  # 如果是扫内网，则线程为5
                else:
                    threadNum = 200  # 扫外网则线程为300

                for num in range(1, threadNum + 1):
                    t = detect(alive_host_queue, winVul_list, proxy)  # 实例化漏洞类，传递参数：存活主机的队列，  存储漏洞的列表
                    threads.append(t)
                    t.start()
                for t in threads:
                    t.join()

        except Exception as e:
            print(r'[-] Load Vul [{}] Error: {}'.format(vulName, e.args))
            continue

    return winVul_list



# 打印漏洞并保存
def printSave_Vul(Vul_list):
    if Vul_list:        # 如果探测出漏洞
        tqdm.write(Fore.BLACK + '-' * 50 + 'Vulnerabilities exist ' + '-' * 50)  # 探测各种漏洞
        for vul in Vul_list:
            Vul_Name, Vul_url, Vul_exist = vul
            tqdm.write(Fore.BLACK + '[{}] {} {}'.format(Vul_Name, Vul_url, Vul_exist))

        # 写入表格里
        vulSheet = saveToExcel(excelSavePath, excel, '漏洞')  # 创建一个ip反查域名页
        vulSheet.saveVul(Vul_list)

    else:
        tqdm.write(Fore.BLACK + '-' * 50 + 'Non-existent vulnerabilities' + '-' * 50)

# 15. 保存相关信息：新的域名和C段IP信息
def saveRelatedInfo(newDomains, ip_count):
    ip2domainSheet = saveToExcel(excelSavePath, excel, '相关域名和C段')    # 创建一个ip反查域名页
    ip2domainSheet.saveNewDomainAndCSubnet(newDomains, ip_count)


# 获取子域名
def run_subdomain():
    # 弃用 1. lijiji
    # Subdomains_ips = lijiejieSubdomain()  # 字典，key为子域名，value为子域名的A记录IP值
    # print(Subdomains_ips)

    # 弃用 2. sublist3r
    # sublist3rSubdomains = []     # Sublist3r()
    # print('sublist3rSubdomains: {}'.format(sublist3rSubdomains))
    # sublist3rSubdomains = []

    # 启用 2. subdomains
    # Subdomains_ips = subdomains3()  # 字典，key为子域名，value为子域名的A记录IP值
    # print('[total: {}] Subdomains3: {}'.format(len(Subdomains_ips), Subdomains_ips))

    # 0. beian2NewDomain
    companyName = beian2NewDomain()

    # 爱企查
    Aiqicha(companyName)

    Subdomains_ips = {}

    # dns域传送
    # subdomains = dnsZoneTransfer()


    # 判断是否是泛解析
    isPanAnalysis = checkPanAnalysis()

    if not isPanAnalysis and ksubdomain:
        # 0. 调用kSubdomain脚本
        ksubdomains = callKsubdomain()
    else:
        ksubdomains = []

    print('[total: {}] ksubdomain: {}'.format(len(ksubdomains), ksubdomains))
    subdomains = printGetNewSubdomains([], ksubdomains)
    print('len [{}]'.format(len(subdomains)))

    # 1. theHarvester
    theHarvesterSubdomains = [] # theHarvester()
    print('[total: {}] theHarvester: {}'.format(len(theHarvesterSubdomains), theHarvesterSubdomains))
    subdomains = printGetNewSubdomains(subdomains, theHarvesterSubdomains)
    print('len [{}]'.format(len(subdomains)))

    # 2. virustotal|ce.baidu.com|www.threatcrowd.org|url.fht.im|qianxun
    othersApiTotalSubdomains = othersApiSubdomain()
    print('[total: {}] webAPI: {}'.format(len(othersApiTotalSubdomains), othersApiTotalSubdomains))
    subdomains = printGetNewSubdomains(subdomains, othersApiTotalSubdomains)
    print('len [{}]'.format(len(subdomains)))

    # 3. github
    githubApiSubdomains = githubApiSubdomain()
    print('[total: {}] Github: {}'.format(len(githubApiSubdomains), githubApiSubdomains))
    subdomains = printGetNewSubdomains(subdomains, githubApiSubdomains)

    # 4. 爬虫(百度｜必应)
    spiderSubdomains = SpiderSubdomain()
    print('[total: {}] Spider: {}'.format(len(spiderSubdomains), spiderSubdomains))
    subdomains = printGetNewSubdomains(subdomains, spiderSubdomains)

    # 防止程序奔溃后重新跑耗费大量时间，所以在当前目录创建文本保存子域名
    with open('{}.txt'.format(domain), 'at') as f:
        for subdomain in subdomains:
            f.writelines('{}\n'.format(subdomain))

    # 测试
    '''
    Subdomains_ips = {}
'''

    # 5. 爬证书
    certsSubdomains = crawlCerts(subdomains)
    print('[total: {}] Certs: {}'.format(len(certsSubdomains), certsSubdomains))
    subdomains = printGetNewSubdomains(subdomains, certsSubdomains)

    # 6. 爬友链
    fcSubdomains = FriendChinsSubdomain(subdomains)
    print('[total: {}] Friends: {}'.format(len(fcSubdomains), fcSubdomains))
    subdomains = printGetNewSubdomains(subdomains, fcSubdomains)

    # 防止程序奔溃后重新跑耗费大量时间，所以在当前目录创建文本保存子域名
    with open('{}.txt'.format(domain), 'at') as f:
        for subdomain in subdomains:
            f.writelines('{}\n'.format(subdomain))


    # 7. 整合子域名，对所有子域名判断是否是CDN,解析A记录，并将所有子域名结果保存到excel里
    Subdomains_ips, CDNSubdomainsDict, notCDNSubdomains = checkCDN_queryA_subdomains(Subdomains_ips, subdomains)

    # host碰撞,censysIPS是censys api得到的解析的IP
    censysIPS = hostCollide(Subdomains_ips)

    # 8. 获取所有子域名的参数链接（存活）
    param_Links = run_ParamLinks()

    # 获取C段的IP
    CIP_List = get_CIP(Subdomains_ips, CDNSubdomainsDict, censysIPS)
    print('C段的IP:{}'.format(CIP_List))

    # 8. 跑C段
    run_cSubnet(CIP_List, Subdomains_ips, notCDNSubdomains, param_Links)

# 获取C段资产
def run_cSubnet(CIP_List, Subdomains_ips, notCDNSubdomains, param_Links):
    print(CIP_List)
    print(Subdomains_ips)
    print(notCDNSubdomains)

    SubdomainAndNotCDNIPs = []  # 子域名和非CDN的IP

    for subdomain in notCDNSubdomains:
        for ip in Subdomains_ips[subdomain]:
            SubdomainAndNotCDNIPs.append(ip)
    SubdomainAndNotCDNIPs = list(set(SubdomainAndNotCDNIPs))
    # 防止IP太多，导致查询次数过多被fofa封
    if len(SubdomainAndNotCDNIPs) > 10:
        SubdomainAndNotCDNIPs = []

    # print(notCDNSubdomainIPs)
    # 8. 调用网络空间引擎,查询根域名和C段IP的资产         webSpace_web_host_port 是Web服务             webSpace_service_host_port  是其他服务
    if domain:            # 跑域名的时候，不跑C段
        webSpace_web_host_port, webSpace_service_host_port = run_webSpace(domain, SubdomainAndNotCDNIPs, [], '')
    else:
        webSpace_web_host_port, webSpace_service_host_port = run_webSpace(domain, [], CIP_List, '')           # 网络空间引擎（fofa、shodan）获取的开放web服务的host（IP/domain）
    # print('webSpace_web_host_port: {}'.format(webSpace_web_host_port))
    # print('webSpace_service_host_port: {}'.format(webSpace_service_host_port))

    for subdomain in Subdomains_ips.keys():
        for ip in Subdomains_ips[subdomain]:
            allTargets_Queue.put(ip)
            allTargets_List.append(ip)


    # ip反查的子域名列表
    ip2domainSubdomains = get_ip2domain()
    print('[total: {}] ip2domainSubdomains: {}'.format(len(ip2domainSubdomains), ip2domainSubdomains))
    print('[ip2domain get new subdomains] [{}]'.format(len(list(set(ip2domainSubdomains)-set(list(Subdomains_ips.keys()))))))

    # 9. 整理开放web服务的host, 存放开放web服务器的ip/domain和port，用来后面的cms识别
    web_host_port, web_ip_list = collation_web_host(Subdomains_ips, webSpace_web_host_port, ip2domainSubdomains)
    print('[total: {}] web_host_port'.format(len(web_host_port)))

    # 10. 对IP进行归属地查询
    ip_address_dict = get_ip_address(web_ip_list)



    # 11. 获取标题, 以及存活的web
    web_Title, alive_Web = run_getWebTitle(web_host_port, ip_address_dict)  # 获取C段资产

    # 不仅仅只信息收集-即跑漏洞
    if justInfoGather == 0:
        webVul_list = detect_webVul(alive_Web)  # 获取C段资产


        if domain:
            # paramVul_list = detect_paramVul(param_Links)      不跑注入
            paramVul_list = []
        else:
            paramVul_list = []

        # 13. 未授权漏洞检测
        unauthWeakVul_list = detect_unauthWeakVul(webSpace_service_host_port)       # 获取C段资产
        # unauthWeakVul_list = []
        # 14. 打印并保存漏洞
        Vul_list = webVul_list + unauthWeakVul_list + paramVul_list
        printSave_Vul(Vul_list)



    # 15. 保存相关信息：新的域名和C段IP信息
    saveRelatedInfo(newDomains, ip_count)

    cprint(r'新的域名：{}'.format(newDomains), 'green')
    cprint(r'C段IP：{}'.format(CIP_List), 'green')
    cprint(r'资产信息保存路径：{}'.format('{}/{}.xlsx'.format(save_fold_path, excel_name)), 'green')
    cprint(r'Github信息保存路径：{}/{}_github.txt'.format(save_fold_path, domain), 'green')

    if domain:
        ret = ""
        for cip in CIP_List:
            ret += cip
            ret += ","
        cprint(r"请使用-c功能跑C段资产", 'green')
        cprint(r"python3 ShuiZe.py -c {}".format(ret[:-1]), 'red')

# 跑fofa Title漏洞
def run_fofaTitle():
    webSpace_web_host_port, webSpace_service_host_port = run_webSpace(domain, [], [], fofaTitle)
    # print('webSpace_web_host_port: {}'.format(webSpace_web_host_port))
    # print('webSpace_service_host_port: {}'.format(webSpace_service_host_port))

    # 整理fofaTitle结果的域名和IP
    collation_fofaDomainIP(webSpace_web_host_port)

    # 9. 整理开放web服务的host, 存放开放web服务器的ip/domain和port，用来后面的cms识别
    web_host_port, web_ip_list = collation_web_host({}, webSpace_web_host_port, [])
    print('[total: {}] web_host_port'.format(len(web_host_port)))

    # 10. 对IP进行归属地查询
    ip_address_dict = get_ip_address(web_ip_list)

    # 11. 获取标题, 以及存活的web
    web_Title, alive_Web = run_getWebTitle(web_host_port, ip_address_dict)

    # 12. Web漏洞检测
    webVul_list = detect_webVul(alive_Web)

    # 13. 检测未授权，弱口令漏洞
    unauthWeakVul_list = detect_unauthWeakVul(webSpace_service_host_port)       # 跑fofa Title漏洞

    # 14. 打印并保存漏洞
    Vul_list = webVul_list + unauthWeakVul_list
    printSave_Vul(Vul_list)

    # 15. 保存相关信息：新的域名和C段IP信息
    saveRelatedInfo(newDomains, ip_count)

    cprint(r'新的域名：{}'.format(newDomains), 'green')
    cprint(r'C段IP：{}'.format(CIP_List), 'green')
    cprint(r'资产信息保存路径：{}'.format('{}/{}.xlsx'.format(save_fold_path, excel_name)), 'green')
    cprint(r'Github信息保存路径：{}/{}_github.txt'.format(save_fold_path, domain), 'green')

# 扫描内网Web漏洞
def run_intranetWeb():
    # 9. 整理开放web服务的host, 存放开放web服务器的ip/domain和port，用来后面的cms识别
    web_host_port, web_ip_list = collation_web_host({}, [], allTargets_List)
    print('[total: {}] web_host_port'.format(len(web_host_port)))

    # 11. 获取标题, 以及存活的web
    web_Title, alive_Web = run_getWebTitle(web_host_port, {})

    # 12. Web漏洞检测
    webVul_list = detect_webVul(alive_Web)

    # 14. 打印并保存漏洞
    Vul_list = webVul_list + []
    printSave_Vul(Vul_list)

    cprint(r'保存路径：{}'.format('{}/{}.xlsx'.format(save_fold_path, excel_name)), 'green')

# 读取文件扫描，文件是每行一个url
def run_file():
    # 9. 整理开放web服务的host, 存放开放web服务器的ip/domain和port，用来后面的cms识别
    web_host_port, web_ip_list = collation_web_host({}, [], allTargets_List)
    print('[total: {}] web_host_port'.format(len(web_host_port)))

    # 10. 对IP进行归属地查询
    ip_address_dict = get_ip_address(web_ip_list)

    # 11. 获取标题, 以及存活的web
    web_Title, alive_Web = run_getWebTitle(web_host_port, ip_address_dict)

    # 12. Web漏洞检测
    webVul_list = detect_webVul(alive_Web)              # # 读取文件扫描

    # 13. 检测未授权，弱口令漏洞
    # webVul_list = []
    # webSpace_service_host_port = []
    unauthWeakVul_list = detect_unauthWeakVul([])       # 读取文件扫描

    # 14. 打印并保存漏洞
    Vul_list = webVul_list + unauthWeakVul_list
    printSave_Vul(Vul_list)

    cprint(r'保存路径：{}'.format('{}/{}.xlsx'.format(save_fold_path, excel_name)), 'green')


# 内网C段扫描， web端口，未授权服务端口，弱口令端口
def run_intranet_cSubnet():
    # 10. 内网端口扫描, 返回存活IP和开放web端口的ip
    web_host_port, alive_host_List = scan_IntranetPorts()
    # web_host_port, alive_host_List = ['http://192.168.168.139:80'], ['192.168.168.139']
    # 11. 获取标题, 以及存活的web
    web_Title, alive_Web = run_getWebTitle(web_host_port, {})

    # 12. Web漏洞检测
    webVul_list = detect_webVul(alive_Web)

    # 13. 不检测未授权，弱口令漏洞，因为需要用proxychains，不能用代理

    # 14. 检测windows漏洞: CVE-2020-0796
    # webVul_list = []
    # alive_host_List = ['192.168.168.148']
    winVul_list = detect_winVul(alive_host_List)

    # 15. 打印并保存漏洞
    printSave_Vul(webVul_list+winVul_list)

    cprint(r'保存路径：{}'.format('{}/{}.xlsx'.format(save_fold_path, excel_name)), 'green')




# 内网服务漏洞检测
def run_intranet_ServiceVul():
    # 从表格里读取《内网服务》的数据，并保存到service_host_port列表里
    intranetServiceSheet = xlsxFileWB.get_sheet_by_name(r'内网服务')
    service_host_port = []
    for i in range(2, intranetServiceSheet.max_row + 1):  # 遍历每行
        eachline = []
        for j in range(1, intranetServiceSheet.max_column + 1):  # 遍历每列
            eachValue = intranetServiceSheet.cell(row=i, column=j).value
            if j == 3:
                eachValue = int(eachValue)
            eachline.append(eachValue)
        service_host_port.append(eachline)      # []


    # 13. 检测未授权，弱口令漏洞
    unauthWeakVul_list = detect_unauthWeakVul(service_host_port)

    # 14. 打印并保存漏洞
    printSave_Vul(unauthWeakVul_list)

    cprint(r'保存路径：{}'.format('{}/{}.xlsx'.format(save_fold_path, excel_name)), 'green')

# 读取masNmap.xlsx文件扫描web漏洞和未授权漏洞
def run_masNmap():
    xlsxFile = openpyxl.load_workbook(masNmapFile)  # 打开文件
    masNmapSheet = xlsxFile.get_sheet_by_name(r'masNmap')
    service_host_port = []
    web_host_port = []
    for i in range(2, masNmapSheet.max_row + 1):  # 遍历每行
        eachline = []
        for j in range(1, masNmapSheet.max_column + 1):  # 遍历每列
            eachValue = masNmapSheet.cell(row=i, column=j).value
            if j == 3:
                eachValue = int(eachValue)
            eachline.append(eachValue)

        if 'http' in eachline[0]:
            url = '{}://{}:{}'.format(eachline[0], eachline[1], eachline[2])
        else:
            url = 'http://{}:{}'.format(eachline[1], eachline[2])
        web_host_port.append(url)
        service_host_port.append(eachline)
        # if 'http' in eachline[0]:
        #     url = '{}://{}:{}'.format(eachline[0], eachline[1], eachline[2])
        #     web_host_port.append(url)
        # else:
        #     service_host_port.append(eachline)

    # 11. 获取标题, 以及存活的web
    web_Title, alive_Web = run_getWebTitle(web_host_port, {})

    # 12. Web漏洞检测
    webVul_list = detect_webVul(alive_Web)

    # 13. 检测未授权，弱口令漏洞
    unauthWeakVul_list = detect_unauthWeakVul(service_host_port)

    # 14. 打印并保存漏洞
    Vul_list = webVul_list + unauthWeakVul_list
    printSave_Vul(Vul_list)

    cprint(r'保存路径：{}'.format('{}/{}.xlsx'.format(save_fold_path, excel_name)), 'green')

def banner():
    banner = '''    __             ____    ___     ____  
   /  \   __ __   |__  |  |_  )   |__  | 
  | () |  \ \ /     / /    / /      / /  
   \__/   /_\_\    /_/    /___|    /_/       author:ske
   
   最好在配置文件里填入fofa、shodan、github、censys的API，这样效果最佳。
   请一定要配置fofa的api～～～最好是高级会员
   配置文件地址：iniFile/config.ini
'''
    print(banner)

# 判断是否是最新版本
def checkVersion():
    with open("versionFlag.txt", "rt", encoding="utf-8") as f:
        now_version = f.read().strip()
    print("目前版本: \n{}".format(now_version))
    version_url = "https://raw.githubusercontent.com/0x727/ShuiZe_0x727/master/versionFlag.txt"
    headers = {'user-agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
    try:
        res = requests.get(url=version_url, headers=headers, timeout=10, verify=False)
        new_version = res.text.strip()
        print("最新版本: \n{}".format(new_version))
        if now_version == new_version:
            cprint("目前版本最新", 'red')
        else:
            add_version = new_version.replace(now_version, "")
            cprint("更新内容如下:{}\n".format(add_version), "red")
            cprint("目前版本非最新，建议及时更新...\n地址: https://github.com/0x727/ShuiZe_0x727/", 'red')

    except Exception as e:
        print('获取版本信息失败...')

# 初始配置
def _init():
    global domain, cSubnet, save_fold_path, excel, excel_name, excelSavePath, proxy, \
        requests_proxies, isIntranet, xlsxFileWB, weak, CIP_List, allTargets_List, \
        allTargets_Queue, masNmapFile, newDomains, ip_count, fofaTitle, ksubdomain, justInfoGather, socksProxysDict

    # python3 %prog -n 1 -c 192.168.1.0,192.168.2.0 -p 1.1.1.1:1111                 内网：使用代理扫描内网C段资产：web标题和漏洞
    # proxychains4 python3 %prog -n 1 -f /result/2ddcaa3ebbd0/172.18.82.0.xlsx      内网：使用proxychains4代理扫描C段的服务漏洞：弱口令和未授权
    # python3 %prog --mn masNmap.xlsx                                               外网：扫描masscan和nmap的结果
    # python3 %prog -n 1 -c 192.168.1.0,192.168.2.0 -v 1                            内网：使用wifi或者vpn的情况下扫web标题和漏洞

    banner()
    checkVersion()

    usage = '\n\t' \
            'python3 %prog -d domain.com\n\t' \
            'python3 %prog -d domain.com --justInfoGather 1\n\t' \
            'python3 %prog -d domain.com --ksubdomain 0\n\t' \
            'python3 %prog -c 192.168.1.0,192.168.2.0,192.168.3.0\n\t' \
            'python3 %prog -f url.txt\n\t' \
            'python3 %prog -n 1 -c 192.168.1.0,192.168.2.0 -p 1.1.1.1:1111\n\t' \
            'python3 %prog -n 1 -f url.txt -p 1.1.1.1:1111 --web 1\n\t' \
            'python3 %prog -n 1 -c 192.168.1.0,192.168.2.0 -v 1\n\t' \
            'proxychains4 python3 %prog -n 1 -f /result/2ddcaa3ebbd0/172.18.82.0.xlsx\n\t' \
            'proxychains4 python3 %prog -n 1 -w 1 -f /result/2ddcaa3ebbd0/172.18.82.0.xlsx\n\t' \
            'python3 %prog --mn masNmap.xlsx\n\t' \
            'python3 %prog --mn masNmap.xlsx -w 1\n\t' \
            'python3 %prog --fofaTitle 大学\n\t' \
            'python3 %prog --domainFile domain.txt\n\t'
    parse = OptionParser(usage=usage)
    parse.add_option('-d', '--domain', dest='domain', type='string', help='target domain')
    parse.add_option('-c', '--cSubnet', dest='cSubnet', type='string', help='target cSubnet')
    parse.add_option('-n', '--intranet', dest='isIntranet', type='int', default=0, help='Scan intranet value set to 1')        # 扫描内网, 值为1扫内网， 默认为0
    parse.add_option('-p', '--proxy', dest='proxy', type='string', default=None, help='Intranet proxy socks5 socks4')           # 代理，socks5和socks4, 默认为None，可用于外网扫描，也可以用于内网扫描
    parse.add_option('-f', '--file', dest='File', type='string', default=None, help='/result/2ddcaa3ebbd0/172.18.82.0.xlsx')  # 扫描内网的服务漏洞-未授权和弱口令
    parse.add_option('-w', '--weak', dest='weak', type='int', default=None, help='run weak password script')                    # 内网弱口令是否要跑
    parse.add_option('-v', '--vpn', dest='vpn', type='int', default=None, help='Run in the case of vpn')            # 在vpn的情况下跑
    parse.add_option('--web', dest='web', type='int', default=None, help='detect web in Intranet')  # 跑内网的web漏洞
    parse.add_option('--mn', dest='masNmapFile', type='str', default=None, help='run masscan nmap result')          # 跑masscan和nmap的结果
    parse.add_option('--fofaTitle', dest='fofaTitle', type='str', default=None, help='run fofa title')  # 跑fofa的title
    parse.add_option('--domainFile', dest='domainFile', type='str', default=None, help='run domain title')  # 跑多个域名
    parse.add_option('--ksubdomain', dest='ksubdomain', type='int', default=1, help='not run ksubdomain')  # 不使用ksubdomain跑子域名
    parse.add_option('--test', dest='testDemo', type='int', default=0, help='if test=1 then run testDemo')  # 测试某个功能
    parse.add_option('--justInfoGather', dest='justInfoGather', type='int', default=0, help='just infoGather, not detect vul')  # 只信息收集，不跑漏洞
    parse.add_option('--getSocks', dest='getSocks', type='int', default=0, help='get socks')  # 获取socks代理

    options, args = parse.parse_args()
    domain, cSubnet, isIntranet, proxy, File, weak, vpn, masNmapFile, fofaTitle, domainFile, web, ksubdomain, justInfoGather, testDemo, getSocks = options.domain, options.cSubnet, options.isIntranet, options.proxy, options.File, options.weak, options.vpn, options.masNmapFile, options.fofaTitle, options.domainFile, options.web, options.ksubdomain, options.justInfoGather, options.testDemo, options.getSocks

    # 所有目标
    allTargets_List = []
    allTargets_Queue = Queue(-1)

    # C段IP列表
    CIP_List = []

    # C段出现的IP个数
    ip_count = Counter()

    # 和目标资产相关联的新的根域名
    newDomains = []

    # 代理
    socksProxysDict = {"baidu": [], "google": []}

    print(domain, cSubnet, isIntranet, proxy, File)

    # requests代理
    if proxy:
        requests_proxies = {"http": "socks5://{}".format(proxy), "https": "socks5://{}".format(proxy)}
    else:
        requests_proxies = None

    # 分割C段，获取ip
    if cSubnet:
        CIP_List = cSubnet.split(',')
        for CIP in CIP_List:
            for ip in IP('{}/24'.format(CIP)):
                allTargets_Queue.put(str(ip))
                allTargets_List.append(str(ip))

    # 扫描外网时加载文件扫描
    if File and not isIntranet:
        with open(File, 'rt') as f:
            for each in f.readlines():
                allTargets_Queue.put(each.strip())
                allTargets_List.append(each.strip())

    # 创建目录
    # 扫描内网漏洞（Web或者服务）
    if File and isIntranet:
        if not web:     # 扫描内网服务漏洞
            save_fold_path, _excel_name = File.rsplit('/', 1)
            excel_name = _excel_name.rsplit('.', 1)[0] + '_ServiceVul'
            xlsxFileWB = openpyxl.load_workbook(File)  # 打开文件
        else:           # 扫描内网web漏洞
            save_fold_path = os.getcwd() + '/result/' + str(uuid4()).split('-')[-1]  # 保存路径
            os.makedirs(save_fold_path)
            with open(File, 'rt') as f:
                for each in f.readlines():
                    allTargets_Queue.put(each.strip())
                    allTargets_List.append(each.strip())

    # 扫描外网或者外网读取file.txt或者读取masNmap.xlsx时
    else:
        try:
            save_fold_path = os.getcwd() + '/result/' + str(uuid4()).split('-')[-1] # 保存路径
            os.makedirs(save_fold_path)
        except Exception:
            pass



    excel = openpyxl.Workbook()
    excel.remove(excel[excel.sheetnames[0]])  # 删除第一个默认的表





    if domain and cSubnet:
        cprint('Error： domain and cSubnet can only pass one', 'red')
        exit(0)
    elif domain and not cSubnet:        # 跑域名
        cprint('-' * 50 + 'Start {} information collection'.format(domain) + '-' * 50, 'green')
        excel_name = domain
        excelSavePath = '{}/{}.xlsx'.format(save_fold_path, excel_name)
        run_subdomain()
    elif not domain and cSubnet:        # 跑C段
        if isIntranet == 0:             # 外网C段
            cprint('-' * 50 + 'Start {} cSubnet collection'.format(cSubnet) + '-' * 50, 'green')
            excel_name = cSubnet
            excelSavePath = '{}/{}.xlsx'.format(save_fold_path, excel_name)
            print('C Subnet: {}'.format(CIP_List))
            run_cSubnet(CIP_List, {}, [], [])
        elif isIntranet == 1:
            if proxy or vpn:            # 内网C段的扫描
                cprint('-' * 50 + 'Start {} cSubnet intranet scan'.format(cSubnet) + '-' * 50, 'green')
                excel_name = cSubnet
                excelSavePath = '{}/{}.xlsx'.format(save_fold_path, excel_name)
                print('C Subnet: {}'.format(CIP_List))
                run_intranet_cSubnet()
            else:
                cprint('Error： Please pass in the agent when scanning the intranet', 'red')

    elif File:
        if isIntranet and not web:              # 扫描内网的服务漏洞
            cprint('-' * 50 + 'Open {} Scanning for service vulnerabilities on the intranet'.format(File) + '-' * 50, 'green')
            excelSavePath = '{}/{}.xlsx'.format(save_fold_path, excel_name)
            print('xlsxFile: {}'.format(File))
            run_intranet_ServiceVul()
        elif isIntranet and web:                   # 扫描内网Web漏洞
            cprint('-' * 50 + 'Open {} Scanning for intranet Web vulnerabilities'.format(File) + '-' * 50, 'green')
            excel_name = str(uuid4()).split('-')[0]
            excelSavePath = '{}/{}.xlsx'.format(save_fold_path, excel_name)
            run_intranetWeb()
        else:                       # 扫描外网漏洞
            cprint('-' * 50 + 'Open {} Scanning'.format(File) + '-' * 50, 'green')
            excel_name = str(uuid4()).split('-')[0]
            excelSavePath = '{}/{}.xlsx'.format(save_fold_path, excel_name)
            print('open File: {}'.format(File))
            run_file()

    elif masNmapFile:   # 跑masscan和nmap的结果
        cprint('-' * 50 + 'Open masNmap File {} to Scanning'.format(masNmapFile) + '-' * 50, 'green')
        excel_name = str(uuid4()).split('-')[0]
        excelSavePath = '{}/{}.xlsx'.format(save_fold_path, excel_name)
        run_masNmap()
    elif fofaTitle:             # 跑fofa Title漏洞
        cprint('-' * 50 + 'Run Fofa Search Title {} to Scanning'.format(fofaTitle) + '-' * 50, 'green')
        excel_name = str(uuid4()).split('-')[0]
        excelSavePath = '{}/{}.xlsx'.format(save_fold_path, excel_name)
        run_fofaTitle()
    elif domainFile:        #  跑域名文件
        cprint('-' * 50 + 'Run Domain File {} to information collection'.format(domainFile) + '-' * 50, 'green')
        with open(domainFile, 'rt') as f:
            for each in f.readlines():
                # C段IP列表
                CIP_List = []
                # C段出现的IP个数
                ip_count = Counter()
                # 和目标资产相关联的新的根域名
                newDomains = []

                domain = each.strip()
                cprint('-' * 50 + 'Start {} information collection'.format(domain) + '-' * 50, 'green')
                excel_name = domain
                excelSavePath = '{}/{}.xlsx'.format(save_fold_path, excel_name)
                excel = openpyxl.Workbook()
                excel.remove(excel[excel.sheetnames[0]])  # 删除第一个默认的表
                run_subdomain()
    elif testDemo == 1:
        # 测试代码
        domain = ''
        save_fold_path = os.getcwd() + '/result/' + str(uuid4()).split('-')[-1]  # 保存路径
        os.makedirs(save_fold_path)
        excel_name = domain
        excelSavePath = '{}/{}.xlsx'.format(save_fold_path, excel_name)

        CIP_List = []
        Subdomains_ips = {}
        notCDNSubdomains = []
        param_Links = []
        run_cSubnet(CIP_List, Subdomains_ips, notCDNSubdomains, param_Links)

    elif getSocks == 1:
        # 从fofa收集代理
        getSocksProxy()

if __name__ == '__main__':
    _init()


