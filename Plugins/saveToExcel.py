from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

class saveToExcel:
    def __init__(self, excelSavePath, excel, title):
        self.excelSavePath = excelSavePath          # excel的保存路径
        self.excel = excel                       # openpyxl.Workbook()的实例话
        self.sheet = self.excel.create_sheet(title=title)   # 创建工作区
        self.Sheet_line = 1               # 表格的行

    # 保存代理列表
    def saveSocksProxys(self, socksProxysDict):
        if self.Sheet_line == 1:
            self.sheet.cell(self.Sheet_line, 1).value = '谷歌代理'
            self.sheet.cell(self.Sheet_line, 2).value = '百度代理'
            self.Sheet_line += 1

        googleProxys = socksProxysDict["google"]
        baiduProxys = socksProxysDict["baidu"]
        if googleProxys != []:
            for proxy in googleProxys:
                self.sheet.cell(self.Sheet_line, 1).value = proxy
                self.Sheet_line += 1

        self.Sheet_line = 2
        if baiduProxys != []:
            for proxy in baiduProxys:
                self.sheet.cell(self.Sheet_line, 2).value = proxy
                self.Sheet_line += 1

        self.excel.save(self.excelSavePath)

    # 保存备案反查顶级域名的结果
    def saveBeianNewDomains(self, beianNewDomains):
        if self.Sheet_line == 1:
            self.sheet.cell(self.Sheet_line, 1).value = '公司名'
            self.sheet.cell(self.Sheet_line, 2).value = '域名'
            self.sheet.cell(self.Sheet_line, 3).value = '备案时间'
            self.Sheet_line += 1

        for _ in beianNewDomains:
            companyName, newDomain, time = _
            self.sheet.cell(self.Sheet_line, 1).value = companyName
            self.sheet.cell(self.Sheet_line, 2).value = newDomain
            self.sheet.cell(self.Sheet_line, 3).value = time
            self.Sheet_line += 1
        self.excel.save(self.excelSavePath)


    # 保存爱企查的结果
    def saveAiqicha(self, selfIcpinfo_infos, invest_infos, holds_infos, branch_infos):
        if self.Sheet_line == 1:
            self.sheet.cell(self.Sheet_line, 1).value = '备案信息'
            self.sheet.cell(self.Sheet_line, 2).value = '网站名称'
            self.sheet.cell(self.Sheet_line, 3).value = '域名'
            self.sheet.cell(self.Sheet_line, 4).value = '备案号'
            self.Sheet_line += 1

        for _ in selfIcpinfo_infos:
            siteName, domain, icpNo = _["siteName"], _["domain"], _["icpNo"]
            self.sheet.cell(self.Sheet_line, 1).value = '备案信息'
            self.sheet.cell(self.Sheet_line, 2).value = siteName
            self.sheet.cell(self.Sheet_line, 3).value = str(domain)
            self.sheet.cell(self.Sheet_line, 4).value = icpNo
            self.Sheet_line += 1






        self.Sheet_line += 1
        self.sheet.cell(self.Sheet_line, 1).value = '对外投资'
        self.sheet.cell(self.Sheet_line, 2).value = '公司名'
        self.sheet.cell(self.Sheet_line, 3).value = '投资占比'
        self.sheet.cell(self.Sheet_line, 4).value = 'pid'
        self.sheet.cell(self.Sheet_line, 5).value = '网站名称'
        self.sheet.cell(self.Sheet_line, 6).value = '域名'
        self.sheet.cell(self.Sheet_line, 7).value = '备案号'
        self.sheet.cell(self.Sheet_line, 8).value = '邮箱地址'
        self.sheet.cell(self.Sheet_line, 9).value = '联系方式'
        self.Sheet_line += 1

        for _ in invest_infos:
            pid, invest_info, icp_info, companyDetail_infos = _["pid"], _["invest_info"], _["icp_info"], _["companyDetail_infos"]
            entName, regRate = invest_info["entName"], invest_info["regRate"]
            emails, telephone = companyDetail_infos["emails"], companyDetail_infos["telephone"]

            if icp_info:
                for each_icp in icp_info:
                    # print(each_icp)
                    siteName, domain, icpNo = each_icp["siteName"], each_icp["domain"], each_icp["icpNo"]
                    self.sheet.cell(self.Sheet_line, 1).value = '对外投资'
                    self.sheet.cell(self.Sheet_line, 2).value = entName
                    self.sheet.cell(self.Sheet_line, 3).value = regRate
                    self.sheet.cell(self.Sheet_line, 4).value = pid
                    self.sheet.cell(self.Sheet_line, 5).value = siteName
                    self.sheet.cell(self.Sheet_line, 6).value = str(domain)
                    self.sheet.cell(self.Sheet_line, 7).value = icpNo
                    self.sheet.cell(self.Sheet_line, 8).value = str(emails)
                    self.sheet.cell(self.Sheet_line, 9).value = str(telephone)
                    self.Sheet_line += 1
            else:
                self.sheet.cell(self.Sheet_line, 1).value = '对外投资'
                self.sheet.cell(self.Sheet_line, 2).value = entName
                self.sheet.cell(self.Sheet_line, 3).value = regRate
                self.sheet.cell(self.Sheet_line, 4).value = pid
                self.sheet.cell(self.Sheet_line, 5).value = ""
                self.sheet.cell(self.Sheet_line, 6).value = ""
                self.sheet.cell(self.Sheet_line, 7).value = ""
                self.sheet.cell(self.Sheet_line, 8).value = str(emails)
                self.sheet.cell(self.Sheet_line, 9).value = str(telephone)
                self.Sheet_line += 1







        self.Sheet_line += 1
        self.sheet.cell(self.Sheet_line, 1).value = '控股企业'
        self.sheet.cell(self.Sheet_line, 2).value = '公司名'
        self.sheet.cell(self.Sheet_line, 3).value = '投资占比'
        self.sheet.cell(self.Sheet_line, 4).value = 'pid'
        self.sheet.cell(self.Sheet_line, 5).value = '网站名称'
        self.sheet.cell(self.Sheet_line, 6).value = '域名'
        self.sheet.cell(self.Sheet_line, 7).value = '备案号'
        self.sheet.cell(self.Sheet_line, 8).value = '邮箱地址'
        self.sheet.cell(self.Sheet_line, 9).value = '联系方式'
        self.Sheet_line += 1


        for _ in holds_infos:
            pid, holds_info, icp_info, companyDetail_infos = _["pid"], _["holds_info"], _["icp_info"], _["companyDetail_infos"]
            entName, proportion = holds_info["entName"], holds_info["proportion"]
            emails, telephone = companyDetail_infos["emails"], companyDetail_infos["telephone"]

            if icp_info:
                for each_icp in icp_info:
                    siteName, domain, icpNo = each_icp["siteName"], each_icp["domain"], each_icp["icpNo"]
                    self.sheet.cell(self.Sheet_line, 1).value = '控股企业'
                    self.sheet.cell(self.Sheet_line, 2).value = entName
                    self.sheet.cell(self.Sheet_line, 3).value = proportion
                    self.sheet.cell(self.Sheet_line, 4).value = pid
                    self.sheet.cell(self.Sheet_line, 5).value = siteName
                    self.sheet.cell(self.Sheet_line, 6).value = str(domain)
                    self.sheet.cell(self.Sheet_line, 7).value = icpNo
                    self.sheet.cell(self.Sheet_line, 8).value = str(emails)
                    self.sheet.cell(self.Sheet_line, 9).value = str(telephone)
                    self.Sheet_line += 1
            else:
                self.sheet.cell(self.Sheet_line, 1).value = '控股企业'
                self.sheet.cell(self.Sheet_line, 2).value = entName
                self.sheet.cell(self.Sheet_line, 3).value = proportion
                self.sheet.cell(self.Sheet_line, 4).value = pid
                self.sheet.cell(self.Sheet_line, 5).value = ""
                self.sheet.cell(self.Sheet_line, 6).value = ""
                self.sheet.cell(self.Sheet_line, 7).value = ""
                self.sheet.cell(self.Sheet_line, 8).value = str(emails)
                self.sheet.cell(self.Sheet_line, 9).value = str(telephone)
                self.Sheet_line += 1







        self.Sheet_line += 1
        self.sheet.cell(self.Sheet_line, 1).value = '分支机构'
        self.sheet.cell(self.Sheet_line, 2).value = '公司名'
        self.sheet.cell(self.Sheet_line, 3).value = 'pid'
        self.sheet.cell(self.Sheet_line, 4).value = '网站名称'
        self.sheet.cell(self.Sheet_line, 5).value = '域名'
        self.sheet.cell(self.Sheet_line, 6).value = '备案号'
        self.sheet.cell(self.Sheet_line, 7).value = '邮箱地址'
        self.sheet.cell(self.Sheet_line, 8).value = '联系方式'
        self.Sheet_line += 1

        for _ in branch_infos:
            pid, branch_info, icp_info, companyDetail_infos = _["pid"], _["branch_info"], _["icp_info"], _["companyDetail_infos"]
            entName = branch_info["entName"]
            emails, telephone = companyDetail_infos["emails"], companyDetail_infos["telephone"]

            if icp_info:
                for each_icp in icp_info:
                    siteName, domain, icpNo = each_icp["siteName"], each_icp["domain"], each_icp["icpNo"]
                    self.sheet.cell(self.Sheet_line, 1).value = '控股企业'
                    self.sheet.cell(self.Sheet_line, 2).value = entName
                    self.sheet.cell(self.Sheet_line, 3).value = pid
                    self.sheet.cell(self.Sheet_line, 4).value = siteName
                    self.sheet.cell(self.Sheet_line, 5).value = str(domain)
                    self.sheet.cell(self.Sheet_line, 6).value = icpNo
                    self.sheet.cell(self.Sheet_line, 7).value = str(emails)
                    self.sheet.cell(self.Sheet_line, 8).value = str(telephone)
                    self.Sheet_line += 1
            else:
                self.sheet.cell(self.Sheet_line, 1).value = '控股企业'
                self.sheet.cell(self.Sheet_line, 2).value = entName
                self.sheet.cell(self.Sheet_line, 3).value = pid
                self.sheet.cell(self.Sheet_line, 4).value = ""
                self.sheet.cell(self.Sheet_line, 5).value = ""
                self.sheet.cell(self.Sheet_line, 6).value = ""
                self.sheet.cell(self.Sheet_line, 7).value = str(emails)
                self.sheet.cell(self.Sheet_line, 8).value = str(telephone)
                self.Sheet_line += 1




        self.excel.save(self.excelSavePath)


    # 保存theHarvester的IP结果
    def saveTheHarvesterIp(self, theHarvesterIp):
        self.sheet.cell(self.Sheet_line, 1).value = 'IP'
        self.Sheet_line += 1

        for ip in theHarvesterIp:
            self.sheet.cell(self.Sheet_line, 1).value = ip
            self.Sheet_line += 1
        self.excel.save(self.excelSavePath)


    # 保存邮箱
    def saveEmails(self, emails, aliveEmails):
        self.sheet.cell(self.Sheet_line, 1).value = '收集的邮箱'
        self.sheet.cell(self.Sheet_line, 2).value = '真实的邮箱'
        self.Sheet_line += 1

        for email in emails:
            self.sheet.cell(self.Sheet_line, 1).value = email
            self.Sheet_line += 1

        self.Sheet_line = 2
        for email in aliveEmails:
            self.sheet.cell(self.Sheet_line, 2).value = email
            self.Sheet_line += 1

        self.excel.save(self.excelSavePath)


    # 保存爬虫结果
    def saveSpider(self, spiderName, links):    # spiderName是搜索引擎的名字，例如百度和必应
        if self.Sheet_line == 1:
            self.sheet.cell(self.Sheet_line, 1).value = '爬虫'
            self.sheet.cell(self.Sheet_line, 2).value = '关键字'
            self.sheet.cell(self.Sheet_line, 3).value = '链接'
            self.sheet.cell(self.Sheet_line, 4).value = '标题'
            self.Sheet_line += 1

        for _ in links:
            each_wd, link, title = _
            self.sheet.cell(self.Sheet_line, 1).value = spiderName
            self.sheet.cell(self.Sheet_line, 2).value = each_wd
            self.sheet.cell(self.Sheet_line, 3).value = link
            try:
                self.sheet.cell(self.Sheet_line, 4).value = title
            except Exception as e:
                self.sheet.cell(self.Sheet_line, 4).value = ''
            self.Sheet_line += 1
        self.excel.save(self.excelSavePath)

    # 保存证书结果
    def saveCert(self, trustedDomainDict):
        self.sheet.cell(self.Sheet_line, 1).value = '子域名'
        self.sheet.cell(self.Sheet_line, 2).value = '证书信任域名'
        self.Sheet_line += 1

        for subdomain in trustedDomainDict:
            certs = trustedDomainDict[subdomain]
            for cert in certs:
                self.sheet.cell(self.Sheet_line, 1).value = subdomain
                self.sheet.cell(self.Sheet_line, 2).value = cert
                self.Sheet_line += 1
        self.excel.save(self.excelSavePath)

    # 保存github敏感信息
    def saveGithub(self, gitSensitiveInfo):
        self.sheet.cell(self.Sheet_line, 1).value = '关键字'
        self.sheet.cell(self.Sheet_line, 2).value = '行数'
        self.sheet.cell(self.Sheet_line, 3).value = '内容'
        self.sheet.cell(self.Sheet_line, 4).value = '作者邮箱'
        self.Sheet_line += 1

        for info in gitSensitiveInfo:
            keyword, line, content, email = info
            self.sheet.cell(self.Sheet_line, 1).value = keyword
            self.sheet.cell(self.Sheet_line, 2).value = line
            try:
                self.sheet.cell(self.Sheet_line, 3).value = content
            except Exception as e:
                self.sheet.cell(self.Sheet_line, 3).value = None            # 可能会报错
            self.sheet.cell(self.Sheet_line, 4).value = email
            self.Sheet_line += 1
        self.excel.save(self.excelSavePath)

    # 保存动态链接和后台地址
    def saveparamHtLinks(self, paramLinks, htLinks):
        self.sheet.cell(self.Sheet_line, 1).value = '链接'
        self.sheet.cell(self.Sheet_line, 2).value = '标题'
        self.Sheet_line += 1

        self.sheet.cell(self.Sheet_line, 1).value = '动态链接'
        self.sheet.cell(self.Sheet_line, 2).value = len(paramLinks)
        self.Sheet_line += 1

        for paramLink in paramLinks:
            self.sheet.cell(self.Sheet_line, 1).value = paramLink
            self.Sheet_line += 1

        self.sheet.cell(self.Sheet_line, 1).value = '后台地址'
        self.sheet.cell(self.Sheet_line, 2).value = len(htLinks)
        self.Sheet_line += 1

        for _ in htLinks:
            link, title = _
            self.sheet.cell(self.Sheet_line, 1).value = link
            self.sheet.cell(self.Sheet_line, 2).value = title
            try:
                self.sheet.cell(self.Sheet_line, 2).value = title
            except Exception as e:
                self.sheet.cell(self.Sheet_line, 2).value = ''
            self.Sheet_line += 1

        self.excel.save(self.excelSavePath)

    # 保存A记录结果
    def saveQueryA(self, Subdomains_ips, CDNSubdomainsDict):
        self.sheet.cell(self.Sheet_line, 1).value = '子域名'
        self.sheet.cell(self.Sheet_line, 2).value = 'A记录IP'
        self.sheet.cell(self.Sheet_line, 3).value = 'CDN'
        self.Sheet_line += 1

        for subdomain in Subdomains_ips:
            ips = str(Subdomains_ips[subdomain])
            self.sheet.cell(self.Sheet_line, 1).value = subdomain
            self.sheet.cell(self.Sheet_line, 2).value = ips
            self.sheet.cell(self.Sheet_line, 3).value = str(CDNSubdomainsDict[subdomain])
            self.Sheet_line += 1
        self.excel.save(self.excelSavePath)

    # 保存host碰撞结果
    def saveHostCollide(self, hostCollideResult):
        self.sheet.cell(self.Sheet_line, 1).value = 'Host'
        self.sheet.cell(self.Sheet_line, 2).value = 'URL'
        self.sheet.cell(self.Sheet_line, 3).value = '状态码'
        self.sheet.cell(self.Sheet_line, 4).value = '带Host的标题'
        self.sheet.cell(self.Sheet_line, 5).value = '不带Host的标题'
        self.Sheet_line += 1

        for _ in hostCollideResult:
            host, ip, code, title, title2 = _
            self.sheet.cell(self.Sheet_line, 1).value = host
            self.sheet.cell(self.Sheet_line, 2).value = ip
            self.sheet.cell(self.Sheet_line, 3).value = code
            self.sheet.cell(self.Sheet_line, 4).value = str(title)
            self.sheet.cell(self.Sheet_line, 5).value = str(title2)
            self.Sheet_line += 1
        self.excel.save(self.excelSavePath)

    # 保存fofa和shodan的结果
    def saveWebSpace(self, webSpaceName, webSpaceResult, query_str):
        if self.Sheet_line == 1:
            self.sheet.cell(self.Sheet_line, 1).value = '空间引擎名'
            self.sheet.cell(self.Sheet_line, 2).value = 'host'
            self.sheet.cell(self.Sheet_line, 3).value = '标题'
            self.sheet.cell(self.Sheet_line, 4).value = 'ip'
            self.sheet.cell(self.Sheet_line, 5).value = '子域名'
            self.sheet.cell(self.Sheet_line, 6).value = '端口'
            self.sheet.cell(self.Sheet_line, 7).value = '服务'
            self.sheet.cell(self.Sheet_line, 8).value = '协议'
            self.sheet.cell(self.Sheet_line, 9).value = '地址'
            self.sheet.cell(self.Sheet_line, 10).value = '查询语句'
            self.sheet.cell(self.Sheet_line, 11).value = 'robots'
            self.sheet.cell(self.Sheet_line, 12).value = '证书'
            self.sheet.cell(self.Sheet_line, 13).value = '公司名'
            self.sheet.cell(self.Sheet_line, 14).value = 'isp'
            self.sheet.cell(self.Sheet_line, 15).value = '收录时间'
            self.Sheet_line += 1

        for result in webSpaceResult:
            if webSpaceName == 'fofa':
                host, title, ip, subdomain, port, server, protocol, address = result
            elif webSpaceName == 'shodan':
                host, title, ip, subdomain, port, server, protocol, address, robots = result
                self.sheet.cell(self.Sheet_line, 11, robots)
            elif webSpaceName == 'quake':
                host, title, ip, subdomain, port, server, protocol, address, cert = result
                self.sheet.cell(self.Sheet_line, 12, str(cert))
            else:
                host, title, ip, subdomain, port, server, protocol, address, company, isp, updated_at = result
                self.sheet.cell(self.Sheet_line, 13, company)
                self.sheet.cell(self.Sheet_line, 14, isp)
                self.sheet.cell(self.Sheet_line, 15, updated_at)

            try:
                title = ILLEGAL_CHARACTERS_RE.sub(r'', title)
            except Exception as e:
                title = ''
            self.sheet.cell(self.Sheet_line, 1).value = webSpaceName
            self.sheet.cell(self.Sheet_line, 2).value = host
            self.sheet.cell(self.Sheet_line, 3).value = title
            self.sheet.cell(self.Sheet_line, 4).value = ip
            self.sheet.cell(self.Sheet_line, 5).value = subdomain
            self.sheet.cell(self.Sheet_line, 6).value = port
            self.sheet.cell(self.Sheet_line, 7).value = server
            self.sheet.cell(self.Sheet_line, 8).value = protocol
            self.sheet.cell(self.Sheet_line, 9).value = address
            self.sheet.cell(self.Sheet_line, 10).value = query_str
            self.Sheet_line += 1
        self.excel.save(self.excelSavePath)

    # 保存网络空间查出来的将非Web服务的结果
    def saveService(self, serviceResult):
        self.sheet.cell(self.Sheet_line, 1).value = '协议'
        self.sheet.cell(self.Sheet_line, 2).value = 'ip'
        self.sheet.cell(self.Sheet_line, 3).value = 'port'
        self.Sheet_line += 1

        for result in serviceResult:
            protocol, ip, port = result
            self.sheet.cell(self.Sheet_line, 1).value = protocol
            self.sheet.cell(self.Sheet_line, 2).value = ip
            self.sheet.cell(self.Sheet_line, 3).value = port
            self.Sheet_line += 1
        self.excel.save(self.excelSavePath)

    # 保存内网存活主机的主机名和所有的IP
    def saveHostNameAndIps(self, alive_hostname_ips):
        self.sheet.cell(self.Sheet_line, 1).value = 'ip'
        self.sheet.cell(self.Sheet_line, 2).value = '主机名'
        self.sheet.cell(self.Sheet_line, 3).value = '其他IP'
        self.Sheet_line += 1

        for each in alive_hostname_ips:
            try:
                if len(each) == 1:
                    self.sheet.cell(self.Sheet_line, 1).value = each[0]
                if len(each) == 2:
                    self.sheet.cell(self.Sheet_line, 1).value = each[0]
                    self.sheet.cell(self.Sheet_line, 2).value = each[1]
                if len(each) > 2:
                    self.sheet.cell(self.Sheet_line, 1).value = each[0]
                    self.sheet.cell(self.Sheet_line, 2).value = each[1]
                    self.sheet.cell(self.Sheet_line, 3).value = each[2]
                    self.Sheet_line += 1
                    for moreIp in each[3:]:
                        self.sheet.cell(self.Sheet_line, 3).value = moreIp
                        self.Sheet_line += 1
            except Exception as e:
                pass

        self.excel.save(self.excelSavePath)

    # 保存IP反查域名结果
    def saveIp2Domain(self, ip2domain_dict):
        self.sheet.cell(self.Sheet_line, 1).value = 'ip'
        self.sheet.cell(self.Sheet_line, 2).value = '域名'
        self.Sheet_line += 1

        for ip in ip2domain_dict.keys():
            ip, subdomains = ip, ip2domain_dict[ip]
            self.sheet.cell(self.Sheet_line, 1).value = ip  # c段的ip
            self.sheet.cell(self.Sheet_line, 2).value = str(subdomains)  # ip反查出来的子域名
            self.Sheet_line += 1
        self.excel.save(self.excelSavePath)

    # 保存存活web的标题，后台
    def saveWebTitle(self, web_Titles):
        self.sheet.cell(self.Sheet_line, 1).value = 'url'
        self.sheet.cell(self.Sheet_line, 2).value = '状态码'
        self.sheet.cell(self.Sheet_line, 3).value = '标题'
        self.sheet.cell(self.Sheet_line, 4).value = 'ip地址'
        self.sheet.cell(self.Sheet_line, 5).value = '框架信息'
        self.sheet.cell(self.Sheet_line, 6).value = '后台路径'
        self.Sheet_line += 1

        for web_Title in web_Titles:
            url, webCode, webTitle, address, info, background = web_Title
            self.sheet.cell(self.Sheet_line, 1).value = url
            self.sheet.cell(self.Sheet_line, 2).value = webCode
            try:
                self.sheet.cell(self.Sheet_line, 3).value = webTitle
            except Exception as e:
                self.sheet.cell(self.Sheet_line, 3).value = None            # 可能会报错
            self.sheet.cell(self.Sheet_line, 4).value = address
            self.sheet.cell(self.Sheet_line, 5).value = info
            try:
                self.sheet.cell(self.Sheet_line, 6).value = background
            except Exception as e:
                self.sheet.cell(self.Sheet_line, 6).value = None
            self.Sheet_line += 1
        self.excel.save(self.excelSavePath)

    # 保存漏洞
    def saveVul(self, Vul_list):
        self.sheet.cell(self.Sheet_line, 1).value = '漏洞名'
        self.sheet.cell(self.Sheet_line, 2).value = 'url'
        self.sheet.cell(self.Sheet_line, 3).value = '状态'
        self.Sheet_line += 1

        for vul in Vul_list:
            Vul_Name, Vul_url, Vul_exist = vul
            self.sheet.cell(self.Sheet_line, 1).value = Vul_Name  # 漏洞名
            self.sheet.cell(self.Sheet_line, 2).value = Vul_url  # 存在漏洞的url
            self.sheet.cell(self.Sheet_line, 3).value = Vul_exist  # 是否存在漏洞,YES存在，NO不存在, Maybe可能
            self.Sheet_line += 1
        self.excel.save(self.excelSavePath)

    # 保存相关域名和C段IP信息
    def saveNewDomainAndCSubnet(self, newDomains, ip_count):
        self.sheet.cell(self.Sheet_line, 1).value = '相关域名'
        self.sheet.cell(self.Sheet_line, 2).value = '相关C段'
        self.sheet.cell(self.Sheet_line, 3).value = '该C段出现的域名个数'
        self.Sheet_line += 1

        for newDomain in newDomains:
            self.sheet.cell(self.Sheet_line, 1).value = newDomain  # 相关域名
            self.Sheet_line += 1

        self.Sheet_line = 2
        for c_subnet in ip_count:
            ip_nums = ip_count[c_subnet]
            self.sheet.cell(self.Sheet_line, 2).value = c_subnet  # 相关C段
            self.sheet.cell(self.Sheet_line, 3).value = ip_nums  # 相关C段
            self.Sheet_line += 1

        self.excel.save(self.excelSavePath)